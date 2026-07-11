"""Export canonical SDD review JSON tables to Excel workbooks."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


DEFAULT_TABLE_BY_SCHEMA = {
    "sdd-review.review-report": "reviewMatrix",
    "sdd-review-security.review-security-report": "sourceRowValidation.rows",
}

EXCEL_SHEET_INVALID_CHARS = set(r"[]:*?/\\")
MAX_EXCEL_SHEET_NAME_LENGTH = 31


class ExportError(Exception):
    """User-correctable export failure."""


def is_scalar(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def load_report(path: Path) -> dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8") as handle:
            report = json.load(handle)
    except FileNotFoundError as exc:
        raise ExportError(f"Input JSON file not found: {path}") from exc
    except PermissionError as exc:
        raise ExportError(f"Cannot read input JSON file: {path}") from exc
    except json.JSONDecodeError as exc:
        raise ExportError(f"Invalid JSON in {path}: {exc.msg}") from exc
    except OSError as exc:
        raise ExportError(f"Cannot read input JSON file: {path}: {exc.strerror or exc.__class__.__name__}") from exc

    if not isinstance(report, dict):
        raise ExportError("Input JSON must contain a top-level object.")
    return report


def resolve_output_path(input_path: Path, output: str | None) -> Path:
    if output:
        return Path(output)
    return input_path.with_suffix(".xlsx")


def resolve_table_name(report: dict[str, Any], requested_table: str | None) -> str:
    if requested_table:
        return requested_table

    schema_name = report.get("schemaName")
    table_name = DEFAULT_TABLE_BY_SCHEMA.get(schema_name)
    if table_name:
        return table_name

    raise ExportError(
        "Unknown schemaName; provide --table with the table name or nested dotted path to export."
    )


def get_table_rows(report: dict[str, Any], table_name: str) -> list[dict[str, Any]]:
    rows = get_value_at_path(report, table_name)
    if not isinstance(rows, list):
        raise ExportError(f"Selected table/path '{table_name}' must be a list of objects.")
    if not all(isinstance(row, dict) for row in rows):
        raise ExportError(f"Selected table/path '{table_name}' must contain only objects.")
    return rows


def get_value_at_path(report: dict[str, Any], table_path: str) -> Any:
    if table_path in report:
        return report[table_path]

    current: Any = report
    for segment in table_path.split("."):
        if not isinstance(current, dict) or segment not in current:
            raise ExportError(
                f"Selected table/path '{table_path}' was not found in the JSON report."
            )
        current = current[segment]
    return current


def make_excel_sheet_name(table_path: str) -> str:
    sanitized = "".join(
        "_" if character in EXCEL_SHEET_INVALID_CHARS else character
        for character in table_path.strip()
    ).strip("'")
    if not sanitized:
        sanitized = "table"
    return sanitized[:MAX_EXCEL_SHEET_NAME_LENGTH]


def extract_scalar_metadata(report: dict[str, Any], table_name: str) -> dict[str, Any]:
    return {
        key: value
        for key, value in report.items()
        if key != table_name and is_scalar(value)
    }


def flatten_value(value: Any) -> Any:
    if is_scalar(value):
        return value
    if isinstance(value, list) and all(is_scalar(item) for item in value):
        return "; ".join("" if item is None else str(item) for item in value)
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def flatten_rows(rows: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    headers: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                headers.append(key)
                seen.add(key)

    flattened = [
        [flatten_value(row.get(header)) for header in headers]
        for row in rows
    ]
    return headers, flattened


def _set_cell_value(cell: Any, value: Any) -> None:
    cell.value = value
    if isinstance(value, str):
        cell.data_type = "s"


def _append_row(ws: Any, values: list[Any], row_number: int) -> None:
    for column_number, value in enumerate(values, start=1):
        _set_cell_value(ws.cell(row=row_number, column=column_number), value)


def format_sheet(ws: Any) -> None:
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.auto_filter.ref = ws.dimensions

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
            if isinstance(cell.value, str) and len(cell.value) > 40:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[column_letter].width = width


def write_workbook(
    metadata: dict[str, Any],
    table_name: str,
    headers: list[str],
    rows: list[list[Any]],
    output_path: Path,
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    _append_row(summary, ["field", "value"], 1)
    for row_number, (key, value) in enumerate(metadata.items(), start=2):
        _append_row(summary, [key, flatten_value(value)], row_number)

    table_sheet = workbook.create_sheet(title=make_excel_sheet_name(table_name))
    _append_row(table_sheet, headers, 1)
    for row_number, row in enumerate(rows, start=2):
        _append_row(table_sheet, row, row_number)

    format_sheet(summary)
    format_sheet(table_sheet)

    try:
        workbook.save(output_path)
    except OSError as exc:
        raise ExportError(f"Cannot write workbook to {output_path}: {exc.strerror or exc.__class__.__name__}") from exc


def export_report(json_path: Path, output: str | None = None, table: str | None = None) -> Path:
    output_path = resolve_output_path(json_path, output)
    report = load_report(json_path)
    table_name = resolve_table_name(report, table)
    table_rows = get_table_rows(report, table_name)
    metadata = extract_scalar_metadata(report, table_name)
    headers, rows = flatten_rows(table_rows)
    write_workbook(metadata, table_name, headers, rows, output_path)
    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Export a canonical SDD review JSON table to an .xlsx workbook."
    )
    parser.add_argument("json_path", type=Path, help="Path to the canonical JSON report.")
    parser.add_argument("--output", help="Optional output .xlsx path.")
    parser.add_argument("--table", help="Optional top-level table name or nested dotted path to export.")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        output_path = export_report(args.json_path, args.output, args.table)
    except ExportError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print(f"Workbook written: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
