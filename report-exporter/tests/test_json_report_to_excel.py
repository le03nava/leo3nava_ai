import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "python"))

from json_report_to_excel import (  # noqa: E402
    DEFAULT_TABLE_BY_SCHEMA,
    ExportError,
    export_report,
    flatten_value,
    get_table_rows,
    make_excel_sheet_name,
    load_report,
    main,
    resolve_output_path,
    resolve_table_name,
)


def write_json(path: Path, payload: object) -> Path:
    path.write_text(json.dumps(payload), encoding="utf-8")
    return path


def base_review_report() -> dict[str, object]:
    return {
        "schemaName": "sdd-review.review-report",
        "schemaVersion": "1.0",
        "changeName": "example-change",
        "verdict": "PASS",
        "ignoredComplex": {"not": "summary"},
        "reviewMatrix": [
            {
                "id": "TD-001",
                "passed": True,
                "score": 3,
                "tags": ["cli", "excel"],
                "details": {"severity": "mandatory", "owner": "apply"},
                "longText": "x" * 80,
                "formulaLike": "=HYPERLINK('https://example.test')",
            },
            {"id": "TD-002", "extra": "second"},
        ],
    }


def source_row_security_report() -> dict[str, object]:
    return {
        "schemaName": "sdd-review-security.review-security-report",
        "schemaVersion": 2,
        "changeName": "source-row-example",
        "verdict": "PASS",
        "sourceRowValidation": {
            "expectedCount": 155,
            "validatedCount": 155,
            "coverageStatus": "complete",
            "exactOnce": True,
            "groupingFields": ["controlDomain", "corporateSection"],
            "rows": [
                {
                    "sourceId": "SRC-001",
                    "corporateSection": "Evidence handling",
                    "controlDomain": "artifact-authority",
                    "repoProfiles": ["python-exporter"],
                    "runtimeSurface": "local-cli",
                    "dataSurface": "canonical-json",
                    "applies": True,
                    "complies": True,
                    "finding": {"status": "pass", "severity": "none"},
                    "evidenceLocation": "python/json_report_to_excel.py",
                },
                {
                    "sourceId": "SRC-002",
                    "corporateSection": "Files and exports",
                    "controlDomain": "generated-artifacts",
                    "repoProfiles": ["python-exporter", "documentation"],
                    "runtimeSurface": "local-cli",
                    "dataSurface": "xlsx-output",
                    "applies": True,
                    "complies": True,
                    "finding": {"status": "pass", "severity": "none"},
                    "evidenceLocation": "python/tests/test_json_report_to_excel.py",
                },
            ],
        },
    }


def test_requirements_are_minimal() -> None:
    requirements = (ROOT / "python" / "requirements.txt").read_text(encoding="utf-8").splitlines()
    normalized = [line.strip() for line in requirements if line.strip()]

    assert normalized == ["openpyxl>=3.1,<4", "pytest"]
    assert not any("pandas" in line.lower() for line in normalized)
    assert not any("win32" in line.lower() or "com" == line.lower() for line in normalized)


def test_default_output_path_derives_from_input() -> None:
    assert resolve_output_path(Path("reports/review-report.json"), None) == Path(
        "reports/review-report.xlsx"
    )
    assert resolve_output_path(Path("reports/review-report.json"), "custom.xlsx") == Path(
        "custom.xlsx"
    )


def test_known_review_schema_exports_review_matrix(tmp_path: Path) -> None:
    source = write_json(tmp_path / "review-report.json", base_review_report())

    output = export_report(source)

    assert output == tmp_path / "review-report.xlsx"
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["summary", "reviewMatrix"]


def test_security_schema_exports_nested_source_row_rows_by_default(tmp_path: Path) -> None:
    source = write_json(tmp_path / "review-security-report.json", source_row_security_report())

    output = export_report(source)

    assert output == tmp_path / "review-security-report.xlsx"
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["summary", "sourceRowValidation.rows"]
    assert DEFAULT_TABLE_BY_SCHEMA["sdd-review-security.review-security-report"] == "sourceRowValidation.rows"
    assert "sdd.review-security-report" not in DEFAULT_TABLE_BY_SCHEMA


def test_compact_only_security_report_fails_without_table_override(tmp_path: Path) -> None:
    source = write_json(
        tmp_path / "security.json",
        {
            "schemaName": "sdd-review-security.review-security-report",
            "schemaVersion": 1,
            "compactControlValidation": {
                "expectedCount": 1,
                "rows": [{"control": "legacy-file-check", "status": "PASS"}],
            },
        },
    )
    output = tmp_path / "security.xlsx"

    with pytest.raises(ExportError, match="sourceRowValidation.rows.*not found"):
        export_report(source, output=str(output))

    assert not output.exists()


def test_unknown_schema_requires_table() -> None:
    with pytest.raises(ExportError, match="--table"):
        resolve_table_name({"schemaName": "future.schema"}, None)


def test_table_override_for_unknown_schema(tmp_path: Path) -> None:
    source = write_json(
        tmp_path / "future.json",
        {"schemaName": "future.schema", "customRows": [{"name": "row"}]},
    )

    output = export_report(source, table="customRows")

    workbook = load_workbook(output)
    assert "customRows" in workbook.sheetnames


def test_nested_table_override_and_excel_safe_sheet_name(tmp_path: Path) -> None:
    source = write_json(
        tmp_path / "future.json",
        {
            "schemaName": "future.schema",
            "artifactExport": {
                "rows": [{"sourceId": "SRC-001", "status": "PASS"}]
            },
        },
    )

    output = export_report(source, table="artifactExport.rows")

    workbook = load_workbook(output)
    assert "artifactExport.rows" in workbook.sheetnames
    assert len("artifactExport.rows") <= 31


def test_sheet_names_are_sanitized_for_excel() -> None:
    assert make_excel_sheet_name("parent.rows") == "parent.rows"
    assert make_excel_sheet_name("bad/path:name?*") == "bad_path_name__"
    assert make_excel_sheet_name("x" * 40) == "x" * 31


def test_missing_table_fails_before_save(tmp_path: Path) -> None:
    source = write_json(tmp_path / "report.json", {"schemaName": "future.schema"})
    output = tmp_path / "report.xlsx"

    with pytest.raises(ExportError, match="missingRows.*not found"):
        export_report(source, output=str(output), table="missingRows")

    assert not output.exists()


def test_missing_nested_table_fails_before_save_and_mentions_path(tmp_path: Path) -> None:
    source = write_json(
        tmp_path / "report.json",
        {"schemaName": "future.schema", "compactControlValidation": {}},
    )
    output = tmp_path / "report.xlsx"

    with pytest.raises(ExportError, match="compactControlValidation.rows.*not found"):
        export_report(source, output=str(output), table="compactControlValidation.rows")

    assert not output.exists()


def test_invalid_json_shapes_fail_closed_without_workbook(tmp_path: Path) -> None:
    malformed = tmp_path / "malformed.json"
    malformed.write_text("{not json", encoding="utf-8")
    array_json = write_json(tmp_path / "array.json", [])
    non_list_table = write_json(tmp_path / "non-list.json", {"rows": {"id": 1}})
    non_object_rows = write_json(tmp_path / "non-object-rows.json", {"rows": ["bad"]})
    nested_non_list_table = write_json(
        tmp_path / "nested-non-list.json",
        {"compactControlValidation": {"rows": {"id": 1}}},
    )

    with pytest.raises(ExportError, match="Invalid JSON"):
        load_report(malformed)
    with pytest.raises(ExportError, match="top-level object"):
        load_report(array_json)
    with pytest.raises(ExportError, match="list of objects"):
        get_table_rows(load_report(non_list_table), "rows")
    with pytest.raises(ExportError, match="only objects"):
        get_table_rows(load_report(non_object_rows), "rows")
    with pytest.raises(ExportError, match="compactControlValidation.rows.*list of objects"):
        get_table_rows(load_report(nested_non_list_table), "compactControlValidation.rows")

    assert not (tmp_path / "malformed.xlsx").exists()


def test_workbook_content_flattening_and_formatting(tmp_path: Path) -> None:
    source = write_json(tmp_path / "review-report.json", base_review_report())

    output = export_report(source)
    workbook = load_workbook(output)
    summary = workbook["summary"]
    table = workbook["reviewMatrix"]

    summary_values = {row[0].value: row[1].value for row in summary.iter_rows(min_row=2)}
    assert summary_values["schemaName"] == "sdd-review.review-report"
    assert summary_values["schemaVersion"] == "1.0"
    assert summary_values["changeName"] == "example-change"
    assert "reviewMatrix" not in summary_values
    assert "ignoredComplex" not in summary_values

    headers = [cell.value for cell in table[1]]
    assert headers == ["id", "passed", "score", "tags", "details", "longText", "formulaLike", "extra"]
    first_row = [cell.value for cell in table[2]]
    assert first_row[0] == "TD-001"
    assert first_row[1] is True
    assert first_row[2] == 3
    assert first_row[3] == "cli; excel"
    assert first_row[4] == '{"owner":"apply","severity":"mandatory"}'
    assert table["G2"].data_type == "s"

    assert table["A1"].font.bold is True
    assert table.auto_filter.ref == table.dimensions
    assert table.freeze_panes == "A2"
    assert table.column_dimensions["A"].width >= 12
    assert table["F2"].alignment.wrap_text is True
    assert summary["A1"].font.bold is True
    assert summary.freeze_panes == "A2"


def test_source_row_workbook_read_back_validates_headers_rows_and_flattening(tmp_path: Path) -> None:
    source = write_json(tmp_path / "review-security-report.json", source_row_security_report())

    output = export_report(source)
    workbook = load_workbook(output)
    table = workbook["sourceRowValidation.rows"]

    headers = [cell.value for cell in table[1]]
    assert headers == [
        "sourceId",
        "corporateSection",
        "controlDomain",
        "repoProfiles",
        "runtimeSurface",
        "dataSurface",
        "applies",
        "complies",
        "finding",
        "evidenceLocation",
    ]
    assert table.max_row == 3
    assert table["A2"].value == "SRC-001"
    assert table["D2"].value == "python-exporter"
    assert table["I2"].value == '{"severity":"none","status":"pass"}'
    assert table["A1"].font.bold is True
    assert table.auto_filter.ref == table.dimensions
    assert table.freeze_panes == "A2"


def test_flatten_value_handles_scalar_lists_and_complex_values() -> None:
    assert flatten_value(["a", "b", 3]) == "a; b; 3"
    assert flatten_value({"b": 2, "a": 1}) == '{"a":1,"b":2}'
    assert flatten_value([{"b": 2, "a": 1}]) == '[{"a":1,"b":2}]'


def test_cli_exit_codes_and_safe_errors(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    source = write_json(
        tmp_path / "sensitive.json",
        {
            "schemaName": "future.schema",
            "secret": "SECRET_TOKEN_SHOULD_NOT_APPEAR",
            "rows": [{"value": "PII_SHOULD_NOT_APPEAR"}],
        },
    )

    exit_code = main([str(source)])
    captured = capsys.readouterr()

    assert exit_code == 1
    assert "--table" in captured.err
    assert "SECRET_TOKEN_SHOULD_NOT_APPEAR" not in captured.err
    assert "PII_SHOULD_NOT_APPEAR" not in captured.err
    assert "Traceback" not in captured.err
    assert not (tmp_path / "sensitive.xlsx").exists()


def test_cli_success_and_unrelated_file_preservation(tmp_path: Path) -> None:
    source = write_json(tmp_path / "review-report.json", base_review_report())
    output = tmp_path / "custom.xlsx"
    sentinel = tmp_path / "sentinel.txt"
    sentinel.write_text("keep", encoding="utf-8")

    exit_code = main([str(source), "--output", str(output)])

    assert exit_code == 0
    assert output.exists()
    assert output.parent == tmp_path
    assert sentinel.read_text(encoding="utf-8") == "keep"


def test_subprocess_has_no_traceback_on_user_error(tmp_path: Path) -> None:
    source = write_json(tmp_path / "future.json", {"schemaName": "future.schema"})

    result = subprocess.run(
        [sys.executable, str(ROOT / "python" / "json_report_to_excel.py"), str(source)],
        text=True,
        capture_output=True,
        check=False,
    )

    assert result.returncode == 1
    assert "--table" in result.stderr
    assert "Traceback" not in result.stderr


def test_source_has_no_markdown_parsing_or_unsupported_dependencies() -> None:
    source = (ROOT / "python" / "json_report_to_excel.py").read_text(encoding="utf-8")
    assert "markdown" not in source.lower()
    assert "pandas" not in source.lower()
    assert "win32com" not in source.lower()


def test_readme_documents_required_usage_and_policies() -> None:
    readme = (ROOT / "python" / "README.md").read_text(encoding="utf-8")

    assert "python -m venv" in readme
    assert "pip install -r python/requirements.txt" in readme
    assert "review-report.json" in readme
    assert "--table" in readme
    assert "openpyxl>=3.1,<4" in readme
    assert "pytest" in readme
    assert "sdd-review-security.review-security-report" in readme
    assert "sourceRowValidation.rows" in readme
    assert "compactControlValidation.rows" not in readme
    assert "nested dotted path" in readme
    assert "sdd.review-security-report" not in readme
    assert "overwrite" in readme.lower()
    assert "derived" in readme.lower()
    assert "python -m pytest python/tests" in readme
    assert "openspec/changes/example-change/review-report.json" in readme
